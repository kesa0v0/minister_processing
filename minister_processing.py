from copy import deepcopy
try:
    import openpyxl as xl 
except:
    print("openpyxl이 설치되지 않았습니다. pip install openpyxl을 통해 설치해주세요.")



class minister:
    def __init__(self, tag, name, position, ideology, traits):
        self.tag = tag
        self.name = name
        self.position = position
        self.ideology = ideology
        self.traits = traits
        

def open_file(excel):
    pos = {"head_of_government" : [],
           "foreign_minister" : [],
           "economy_minister" : [],
           "security_minister" : [],
           "intel_minister" : [],
           "defence_minister" : [],
           
           "army_chief" : [],
           "navy_chief" : [],
           "air_chief" : []}
    
    excel = xl.load_workbook(excel)
    ministers = {sheet : deepcopy(pos) for sheet in excel.sheetnames}
    for ws in excel.sheetnames:
        print("loading...: " + ws)
        for row in excel[ws].rows:
            ministers[ws][row[1].value].append(minister(ws, row[0].value, row[1].value, row[2].value, [trait.value for trait in row[3:]]))
            print("is read : " + row[0].value)
        print()
    return ministers

def minister_writedown(minister):
    if minister.ideology == "none":
        available = ""
    else:
        available = f'''
        available = {'{'}
            if = {'{'}
                has_government = {minister.ideology}
            {'}'}
        {'}'}
        '''
    return f"""    {minister.name} = {'{'}
        
        allowed = {'{'}
            original_tag = {minister.tag}
        {'}'}
        
        {available}
        
        traits = {'{'}
            {'''
            '''.join(minister.traits)}
        {'}'}
    {'}'}
    """

def export(ministers):
    for tag, ministers in ministers.items():
        print(f"Processing TAG: {tag}")
        with open(tag + ".txt", 'w', encoding="UTF-8") as f:
            for pos, ministers in ministers.items():
                f.writelines(f"{pos} = {'{'}")
                f.writelines("\n")
                
                for minister in ministers:
                    print(f"Processing name : {minister.name}")
                    f.writelines(minister_writedown(minister))
                
                f.writelines('}\n\n')
        
        print()

try:
    filename = input("처리할 파일의 주소를 입력해 주세요: ")

    ministers = open_file(filename)
    print()
    export(ministers)
    input("아무키나 눌러주세요...")
except Exception as e:
    with open("errorlog.txt", 'w', encoding='UTF-8') as f:
        f.write(type(e), '\n\n\n\n')
        f.write(e.args)